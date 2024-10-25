from flask import Flask, render_template, send_file
from openpyxl import load_workbook
from docxtpl import DocxTemplate
import os
import webbrowser
import json

app = Flask(__name__)
app.config['GENERATED_FOLDER'] = 'generated'  # Carpeta para los archivos generados
app.config['SECRET_KEY'] = '39QDt7fVWUuPqLsPDAF3XkuDQEKiZkxN9z'

# Página principal para generar el informe
@app.route('/')
def index():
    return render_template('index.html')

# Función para formatear valores según el formato de la celda en Excel
def formatear_valor(celda):
    if isinstance(celda.value, (int, float)):
        if celda.value >= 1:
            return "${:,.2f}".format(celda.value)  # Formato de dólares para valores positivos
        elif celda.value < 0:
            return "-${:,.2f}".format(abs(celda.value))  # Formato de dólares para valores negativos
        elif celda.value >= 0:
            return "{:.2%}".format(celda.value)  # Formato de porcentaje para valores menores que 1
    else:
        return celda.value  # No aplicar formato si no es numérico

# Procesar texto, imágenes y generar archivo .docx
@app.route('/generate', methods=['POST'])
def generate_doc():
    # Cargar la plantilla .docx
    doc = DocxTemplate('1.docx')
    context = {}

    # Cargar los archivos Excel directamente desde la raíz del proyecto
    uploaded_files = ['ejemplo1.xlsx', 'ejemplo2.xlsx']  # Especifica aquí tus archivos Excel
    for idx, filename in enumerate(uploaded_files):
        file_path = os.path.join(os.getcwd(), filename)  # Cambia el directorio si es necesario

        if filename.endswith('.xlsx'):
            wb = load_workbook(file_path, data_only=True)

            # Seleccionar la hoja correspondiente
            if 'ejemplo1' in filename:
                ws = wb.worksheets[1]  # Hoja 2
                evidencia_1 = formatear_valor(ws['D9'])  # Celda D9 con formato
                evidencia_2 = formatear_valor(ws['D10'])  # Celda D10 con formato
                context[f'ev1'] = evidencia_1
                context[f'ev2'] = evidencia_2

            elif 'ejemplo2' in filename:
                ws = wb.worksheets[2]  # Hoja 4
                evidencia_3 = formatear_valor(ws['B12'])  # Celda C11 con formato
                ws = wb.worksheets[3]  # Hoja 4
                evidencia_4 = formatear_valor(ws['C13'])  # Celda C12 con formato
                ws = wb.worksheets[5]  # Hoja 4
                evidencia_5 = formatear_valor(ws['C13'])
                evidencia_6 = formatear_valor(ws['K7'])
                evidencia_7 = formatear_valor(ws['K6'])
                evidencia_8 = formatear_valor(ws['F9'])

                context[f'ev3'] = evidencia_3
                context[f'ev4'] = evidencia_4
                context[f'ev5'] = evidencia_5
                context[f'ev6'] = evidencia_6
                context[f'ev7'] = evidencia_7
                context[f'ev8'] = evidencia_8

            # Si no es un archivo reconocido, usa la celda A1 por defecto
            else:
                evidencia_default = formatear_valor(ws['A1'])  # Celda A1 con formato
                context[f'evidencia_default_{idx}'] = evidencia_default

    # Verificar el contenido del contexto para asegurarse de que todo esté correcto
    print("Contexto final:", context)

    # Rellenar la plantilla con los datos
    doc.render(context)

    # Guardar el nuevo archivo .docx lleno
    output_path = os.path.join(app.config['GENERATED_FOLDER'], 'informe_lleno.docx')

    # Eliminar archivo si ya existe
    if os.path.exists(output_path):
        os.remove(output_path)

    doc.save(output_path)

    # Enviar el archivo generado al cliente para su descarga
    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    webbrowser.open("http://127.0.0.1:5000")
    if not os.path.exists(app.config['GENERATED_FOLDER']):
        os.makedirs(app.config['GENERATED_FOLDER'])
    app.run(debug=True, use_reloader=False, threaded=False)
